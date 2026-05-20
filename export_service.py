from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm
from io import BytesIO
from datetime import datetime

class ExportService:
    @staticmethod
    def setup_font():
        """註冊中文字體 (使用 CID 字體 STSong-Light，不需字體檔)"""
        try:
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            return 'STSong-Light'
        except:
            return 'Helvetica'

    @staticmethod
    def generate_pdf(reports, export_type: str):
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        font_name = ExportService.setup_font()
        
        if export_type == 'report':
            ExportService._draw_report_pdf(c, reports, font_name, width, height)
        else:
            ExportService._draw_replacement_pdf(c, reports, font_name, width, height)
            
        c.save()
        buffer.seek(0)
        return buffer

    @staticmethod
    def _draw_report_pdf(c, reports, font, width, height):
        """通報紀錄：一頁一筆"""
        for r in reports:
            # 標題
            c.setFont(font, 18)
            c.drawCentredString(width/2, height - 2*cm, f"公車維修通報紀錄單")
            
            # 基本資訊區
            c.setFont(font, 12)
            y = height - 4*cm
            c.drawString(2*cm, y, f"客運公司：{r.get('vendor_name', '-')}")
            c.drawString(10*cm, y, f"車號：{r.get('car_number', '-')}")
            
            y -= 0.8*cm
            c.drawString(2*cm, y, f"完成時間：{r.get('completed_at', '-')[:16].replace('T', ' ')}")
            c.drawString(10*cm, y, f"處理人員：{r.get('handler_name', '-')}")

            y -= 0.8*cm
            c.drawString(2*cm, y, f"目前里程：{r.get('mileage', '-') or '-'} KM")
            
            # 分隔線
            y -= 0.5*cm
            c.line(2*cm, y, width - 2*cm, y)
            
            # 問題描述
            y -= 1*cm
            c.setFont(font, 14)
            c.drawString(2*cm, y, "【問題描述】")
            y -= 0.8*cm
            c.setFont(font, 11)
            # 簡單處理換行 (ReportLab 不會自動換行，這裡先做截斷或簡單換行)
            desc = r.get('description', '')
            text_obj = c.beginText(2.5*cm, y)
            text_obj.setFont(font, 11)
            for line in ExportService._wrap_text(desc, 40):
                text_obj.textLine(line)
            c.drawText(text_obj)
            
            # 處理方案
            y -= 4*cm
            c.setFont(font, 14)
            c.drawString(2*cm, y, "【處理方案】")
            y -= 0.8*cm
            c.setFont(font, 11)
            sol = r.get('solution', '')
            text_obj = c.beginText(2.5*cm, y)
            text_obj.setFont(font, 11)
            for line in ExportService._wrap_text(sol, 40):
                text_obj.textLine(line)
            c.drawText(text_obj)
            
            # 底部空白預留線
            y = 8*cm
            c.setDash(1, 2)
            c.line(2*cm, y, width - 2*cm, y)
            c.setFont(font, 10)
            c.drawString(2*cm, y - 0.5*cm, "( 以下為預留空白處，供新增資訊或核章使用 )")
            
            c.showPage() # 換頁

    @staticmethod
    def _draw_replacement_pdf(c, reports, font, width, height):
        """換件紀錄：清單式"""
        c.setFont(font, 18)
        c.drawCentredString(width/2, height - 2*cm, f"公車零件更換紀錄表")
        
        y = height - 4*cm
        c.setFont(font, 10)
        # 表格標題
        headers = ["日期", "客運", "車號", "更換零件", "里程", "人員"]
        widths = [3*cm, 3*cm, 3*cm, 6*cm, 2.5*cm, 2*cm]
        
        # 畫表頭
        x = 1*cm
        for i, h in enumerate(headers):
            c.drawString(x, y, h)
            x += widths[i]
        
        y -= 0.3*cm
        c.line(1*cm, y, width - 1*cm, y)
        y -= 0.6*cm
        
        for r in reports:
            if y < 2*cm: # 換頁檢查
                c.showPage()
                y = height - 2*cm
                c.setFont(font, 10)
            
            x = 1*cm
            date_str = r.get('completed_at', '')[:10]
            c.drawString(x, y, date_str)
            x += widths[0]
            c.drawString(x, y, r.get('vendor_name', '-')[:6])
            x += widths[1]
            c.drawString(x, y, r.get('car_number', '-'))
            x += widths[2]
            c.drawString(x, y, r.get('solution', '-')[:15])
            x += widths[3]
            c.drawString(x, y, str(r.get('mileage', '-') or '-'))
            x += widths[4]
            c.drawString(x, y, r.get('handler_name', '-')[:5])
            
            y -= 0.8*cm
            c.setDash(1, 1)
            c.line(1*cm, y + 0.2*cm, width - 1*cm, y + 0.2*cm)
            c.setDash(1, 0) # 恢復實線

    @staticmethod
    def _wrap_text(text, limit):
        """簡單的文字換行處理"""
        if not text: return ["-"]
        lines = []
        for i in range(0, len(text), limit):
            lines.append(text[i:i+limit])
        return lines

    @staticmethod
    def generate_excel(reports):
        """產生 Excel XLSX 檔案，若未安裝 openpyxl 則自動降級為帶有 UTF-8 BOM 的 CSV 檔案"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "維修通報紀錄"
            
            font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="微軟正黑體", size=10)
            fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            headers = ["完成日期", "客運商", "車號", "類型", "里程(KM)", "處理方案", "處理人員", "問題描述"]
            ws.append(headers)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for r in reports:
                completed_at = r.get('completed_at', '')
                if completed_at:
                    completed_at = completed_at[:16].replace('T', ' ')
                
                row_data = [
                    completed_at,
                    r.get('vendor_name', '-'),
                    r.get('car_number', '-'),
                    r.get('solution_type', '-'),
                    r.get('mileage', '-') or '-',
                    r.get('solution', '-'),
                    r.get('handler_name', '-'),
                    r.get('description', '-')
                ]
                ws.append(row_data)
                
            for row_idx in range(2, len(reports) + 2):
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = thin_border
                    if col_idx in [1, 2, 3, 4, 5, 7]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    val_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                    if val_len > max_len:
                        max_len = val_len
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        except ImportError:
            # Fallback to UTF-8 BOM CSV
            import csv
            from io import StringIO
            
            buffer = BytesIO()
            buffer.write(b'\xef\xbb\xbf') # BOM for Excel
            
            text_buffer = StringIO()
            writer = csv.writer(text_buffer)
            writer.writerow(["完成日期", "客運商", "車號", "類型", "里程(KM)", "處理方案", "處理人員", "問題描述"])
            
            for r in reports:
                completed_at = r.get('completed_at', '')
                if completed_at:
                    completed_at = completed_at[:16].replace('T', ' ')
                writer.writerow([
                    completed_at,
                    r.get('vendor_name', '-'),
                    r.get('car_number', '-'),
                    r.get('solution_type', '-'),
                    r.get('mileage', '-') or '-',
                    r.get('solution', '-'),
                    r.get('handler_name', '-'),
                    r.get('description', '-')
                ])
                
            buffer.write(text_buffer.getvalue().encode('utf-8'))
            buffer.seek(0)
            return buffer

